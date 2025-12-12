"""
SISTEMA DE REVISIÓN AUTOMÁTICA DE FECHAS EN EXCEL
Versión para GitHub Actions
"""

import openpyxl
from datetime import datetime, date
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os
import sys

# Obtener configuración desde variables de entorno (GitHub Secrets)
GMAIL_USUARIO = "efasoftt@gmail.com"
GMAIL_PASSWORD = "becwcnkgqlomjgcv"  # Tu contraseña de aplicación de Gmail
EMAIL_DESTINO = "efasoft@hotmail.com"

# Archivo Excel (se descarga automáticamente desde Google Drive)
RUTA_EXCEL = "medicamentos.xlsx"

# Configuración de revisión
COLUMNAS_REVISAR = ['I']
DIAS_ALERTA = 3

def log(mensaje):
    """Registrar mensajes con timestamp"""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{timestamp}] {mensaje}")

def leer_excel_y_buscar_alertas(ruta_archivo):
    """Lee el archivo Excel y busca fechas próximas"""
    try:
        log(f"Abriendo archivo Excel: {ruta_archivo}")
        workbook = openpyxl.load_workbook(ruta_archivo, data_only=True)
        sheet = workbook.active
        
        alertas = []
        fecha_hoy = date.today()
        
        columnas_numeros = [openpyxl.utils.cell.column_index_from_string(col) for col in COLUMNAS_REVISAR]
        
        log(f"Revisando columnas: {', '.join(COLUMNAS_REVISAR)}")
        log(f"Buscando fechas con menos de {DIAS_ALERTA} días...")
        
        for fila in range(1, sheet.max_row + 1):
            for col_num in columnas_numeros:
                celda = sheet.cell(row=fila, column=col_num)
                valor = celda.value
                
                if isinstance(valor, datetime):
                    fecha_celda = valor.date()
                    dias_restantes = (fecha_celda - fecha_hoy).days
                    
                    if 0 <= dias_restantes < DIAS_ALERTA:
                        col_letra = openpyxl.utils.cell.get_column_letter(col_num)
                        info_adicional = sheet.cell(row=fila, column=1).value or ""
                        
                        alerta = {
                            'fila': fila,
                            'columna': col_letra,
                            'fecha': fecha_celda,
                            'dias_restantes': dias_restantes,
                            'informacion': str(info_adicional)
                        }
                        alertas.append(alerta)
                        log(f"  ⚠️ Alerta: Fila {fila}, Columna {col_letra}, Fecha: {fecha_celda}, Días: {dias_restantes}")
        
        workbook.close()
        log(f"Total de alertas encontradas: {len(alertas)}")
        return alertas
    
    except FileNotFoundError:
        log(f"❌ ERROR: No se encontró el archivo: {ruta_archivo}")
        return None
    except Exception as e:
        log(f"❌ ERROR al leer Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def crear_html_email(alertas):
    """Crea el contenido HTML del email"""
    num_alertas = len(alertas)
    fecha_revision = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            body {{ font-family: Arial, sans-serif; padding: 20px; }}
            h2 {{ color: #4472C4; }}
            .alerta {{ background-color: #FFF3CD; padding: 10px; border-left: 4px solid #FFC107; margin: 10px 0; }}
            table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
            th {{ background-color: #4472C4; color: white; padding: 10px; text-align: center; }}
            td {{ padding: 8px; text-align: center; border: 1px solid #ddd; }}
            .hoy {{ background-color: #FFEBEE; }}
            .manana {{ background-color: #FFF3E0; }}
            .proximo {{ background-color: #FFFDE7; }}
            .estado-hoy {{ background-color: #F44336; color: white; font-weight: bold; padding: 5px; }}
            .estado-manana {{ background-color: #FF9800; font-weight: bold; padding: 5px; }}
            .estado-proximo {{ background-color: #FFEB3B; font-weight: bold; padding: 5px; }}
            .footer {{ font-size: 11px; color: #666; margin-top: 30px; }}
        </style>
    </head>
    <body>
        <h2>📅 Reporte Automático de Fechas Próximas</h2>
        <p>Se han detectado <strong style="color: red; font-size: 18px;">{num_alertas} fechas</strong> que vencen en menos de {DIAS_ALERTA} días.</p>
        
        <div class="alerta">
            ⚠️ <strong>Acción requerida:</strong> Revise las fechas marcadas y tome las acciones necesarias.
        </div>
        
        <table>
            <tr>
                <th>Fila</th>
                <th>Columna</th>
                <th>Fecha</th>
                <th>Días Restantes</th>
                <th>Estado</th>
                <th>Información</th>
            </tr>
    """
    
    for alerta in alertas:
        dias = alerta['dias_restantes']
        fecha_formateada = alerta['fecha'].strftime("%d/%m/%Y")
        
        if dias == 0:
            clase_fila = "hoy"
            estado = '<span class="estado-hoy">🔴 HOY</span>'
        elif dias == 1:
            clase_fila = "manana"
            estado = '<span class="estado-manana">🟠 MAÑANA</span>'
        else:
            clase_fila = "proximo"
            estado = '<span class="estado-proximo">🟡 PRÓXIMO</span>'
        
        html += f"""
            <tr class="{clase_fila}">
                <td>{alerta['fila']}</td>
                <td>{alerta['columna']}</td>
                <td>{fecha_formateada}</td>
                <td><strong style="font-size: 16px;">{dias}</strong></td>
                <td>{estado}</td>
                <td>{alerta['informacion']}</td>
            </tr>
        """
    
    html += f"""
        </table>
        
        <div class="footer">
            <hr>
            <p>🕐 Revisión automática realizada: {fecha_revision}</p>
            <p>🤖 Email enviado automáticamente desde: {GMAIL_USUARIO}</p>
            <p>☁️ Sistema ejecutado en GitHub Actions (Cloud)</p>
            <p>📎 El archivo Excel completo se adjunta a este correo.</p>
        </div>
    </body>
    </html>
    """
    
    return html

def enviar_email(destinatario, asunto, cuerpo_html, archivo_adjunto=None):
    """Envía un email vía Gmail SMTP"""
    try:
        log("Preparando email...")
        
        mensaje = MIMEMultipart()
        mensaje['From'] = GMAIL_USUARIO
        mensaje['To'] = destinatario
        mensaje['Subject'] = asunto
        
        mensaje.attach(MIMEText(cuerpo_html, 'html', 'utf-8'))
        
        if archivo_adjunto and os.path.exists(archivo_adjunto):
            log(f"Adjuntando archivo: {archivo_adjunto}")
            with open(archivo_adjunto, 'rb') as archivo:
                parte = MIMEBase('application', 'octet-stream')
                parte.set_payload(archivo.read())
                encoders.encode_base64(parte)
                nombre_archivo = os.path.basename(archivo_adjunto)
                parte.add_header('Content-Disposition', f'attachment; filename= {nombre_archivo}')
                mensaje.attach(parte)
        
        log("Conectando con Gmail SMTP...")
        servidor = smtplib.SMTP('smtp.gmail.com', 587)
        servidor.starttls()
        
        log("Autenticando...")
        servidor.login(GMAIL_USUARIO, GMAIL_PASSWORD)
        
        log("Enviando email...")
        texto = mensaje.as_string()
        servidor.sendmail(GMAIL_USUARIO, destinatario, texto)
        servidor.quit()
        
        log("✅ Email enviado exitosamente!")
        return True
    
    except smtplib.SMTPAuthenticationError:
        log("❌ ERROR DE AUTENTICACIÓN: Usuario o contraseña incorrectos")
        return False
    except Exception as e:
        log(f"❌ ERROR al enviar email: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """Función principal"""
    log("="*70)
    log("SISTEMA DE REVISIÓN AUTOMÁTICA - GITHUB ACTIONS")
    log("="*70)
    
    # Verificar que las variables de entorno estén configuradas
    if not all([GMAIL_USUARIO, GMAIL_PASSWORD, EMAIL_DESTINO]):
        log("❌ ERROR: Faltan variables de entorno (GitHub Secrets)")
        log("   Configura: GMAIL_USUARIO, GMAIL_PASSWORD, EMAIL_DESTINO")
        sys.exit(1)
    
    # Verificar que el archivo Excel existe
    if not os.path.exists(RUTA_EXCEL):
        log(f"❌ ERROR: No se encontró el archivo Excel: {RUTA_EXCEL}")
        log("   Verifica que se descargó correctamente desde Google Drive")
        sys.exit(1)
    
    # Buscar alertas
    alertas = leer_excel_y_buscar_alertas(RUTA_EXCEL)
    
    if alertas is None:
        log("❌ No se pudo leer el archivo Excel")
        sys.exit(1)
    
    # Si hay alertas, enviar email
    if len(alertas) > 0:
        log(f"\n🚨 Se encontraron {len(alertas)} alertas. Preparando email...")
        
        cuerpo_html = crear_html_email(alertas)
        asunto = f"⚠️ ALERTAS: {len(alertas)} Fechas Próximas - {datetime.now().strftime('%d/%m/%Y')}"
        
        resultado = enviar_email(EMAIL_DESTINO, asunto, cuerpo_html, RUTA_EXCEL)
        
        if resultado:
            log(f"✅ Proceso completado exitosamente")
            log(f"📧 Email enviado a: {EMAIL_DESTINO}")
        else:
            log("❌ El email no pudo ser enviado")
            sys.exit(1)
    else:
        log("✅ No se encontraron alertas. No se envió ningún email.")
    
    log("="*70)
    log("PROCESO FINALIZADO")
    log("="*70)

if __name__ == "__main__":
    main()
