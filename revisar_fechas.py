"""
SISTEMA DE REVISIÓN AUTOMÁTICA DE FECHAS EN EXCEL
Versión Mejorada con Bootstrap 5 y WhatsApp
"""

import openpyxl
from datetime import datetime, date
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os
import sys
import requests
import json

# Obtener configuración desde variables de entorno
GMAIL_USUARIO = os.environ.get('GMAIL_USUARIO')
GMAIL_PASSWORD = os.environ.get('GMAIL_PASSWORD')
EMAIL_DESTINO = os.environ.get('EMAIL_DESTINO')
WHATSAPP_API_KEY = os.environ.get('WHATSAPP_API_KEY', '')  # API de CallMeBot (gratis)

# Archivo Excel
RUTA_EXCEL = "medicamentos.xlsx"

# Configuración
COLUMNAS_REVISAR = ['I']
DIAS_ALERTA = 5
FILA_INICIO = 14  # Empezar desde la fila 14

def log(mensaje):
    """Registrar mensajes con timestamp"""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{timestamp}] {mensaje}")

def leer_info_paciente(sheet):
    """Lee la información del paciente desde las celdas específicas"""
    try:
        paciente = sheet['B2'].value or "No especificado"
        ubicacion = sheet['B3'].value or "No especificada"
        telefono_whatsapp = sheet['I4'].value or ""
        
        # Limpiar número de WhatsApp (quitar espacios, guiones, etc.)
        if telefono_whatsapp:
            telefono_whatsapp = str(telefono_whatsapp).replace(" ", "").replace("-", "").replace("+", "")
        
        return {
            'paciente': paciente,
            'ubicacion': ubicacion,
            'telefono': telefono_whatsapp
        }
    except Exception as e:
        log(f"Error al leer información del paciente: {e}")
        return {
            'paciente': "No especificado",
            'ubicacion': "No especificada",
            'telefono': ""
        }

def leer_excel_y_buscar_alertas(ruta_archivo):
    """Lee el archivo Excel y busca fechas próximas desde la fila 14"""
    try:
        log(f"Abriendo archivo Excel: {ruta_archivo}")
        workbook = openpyxl.load_workbook(ruta_archivo, data_only=True)
        sheet = workbook.active
        
        # Leer información del paciente
        info_paciente = leer_info_paciente(sheet)
        log(f"Paciente: {info_paciente['paciente']}")
        log(f"Ubicación: {info_paciente['ubicacion']}")
        
        alertas = []
        fecha_hoy = date.today()
        
        columnas_numeros = [openpyxl.utils.cell.column_index_from_string(col) for col in COLUMNAS_REVISAR]
        
        log(f"Revisando columnas: {', '.join(COLUMNAS_REVISAR)} desde fila {FILA_INICIO}")
        log(f"Buscando fechas con menos de {DIAS_ALERTA} días...")
        
        # Recorrer desde la fila 14 en adelante
        for fila in range(FILA_INICIO, sheet.max_row + 1):
            for col_num in columnas_numeros:
                celda = sheet.cell(row=fila, column=col_num)
                valor = celda.value
                
                if isinstance(valor, datetime):
                    fecha_celda = valor.date()
                    dias_restantes = (fecha_celda - fecha_hoy).days
                    
                    if 0 <= dias_restantes <= DIAS_ALERTA:
                        col_letra = openpyxl.utils.cell.get_column_letter(col_num)
                        
                        # Obtener nombre del medicamento (columna A)
                        nombre_medicamento = sheet.cell(row=fila, column=1).value or "Medicamento sin nombre"
                        
                        # Obtener uso del medicamento (columna B)
                        uso_medicamento = sheet.cell(row=fila, column=2).value or "Uso no especificado"
                        
                        alerta = {
                            'fila': fila,
                            'columna': col_letra,
                            'fecha': fecha_celda,
                            'dias_restantes': dias_restantes,
                            'medicamento': str(nombre_medicamento),
                            'uso': str(uso_medicamento)
                        }
                        alertas.append(alerta)
                        log(f"  ⚠️ Alerta: {nombre_medicamento} - Fila {fila}, Columna {col_letra}, Fecha: {fecha_celda}, Días: {dias_restantes}")
        
        workbook.close()
        log(f"Total de alertas encontradas: {len(alertas)}")
        return alertas, info_paciente
    
    except FileNotFoundError:
        log(f"❌ ERROR: No se encontró el archivo: {ruta_archivo}")
        return None, None
    except Exception as e:
        log(f"❌ ERROR al leer Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return None, None

def crear_html_email_bootstrap(alertas, info_paciente):
    """Crea un email con diseño Bootstrap 5 moderno"""
    num_alertas = len(alertas)
    fecha_revision = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    
    # Agrupar alertas por urgencia
    alertas_hoy = [a for a in alertas if a['dias_restantes'] == 0]
    alertas_manana = [a for a in alertas if a['dias_restantes'] == 1]
    alertas_proximas = [a for a in alertas if a['dias_restantes'] >= 2]
    
    html = f"""
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Alertas de Medicamentos</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <style>
        * {{
            box-sizing: border-box;
        }}
        body {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            padding: 10px;
            margin: 0;
        }}
        .container-email {{
            max-width: 800px;
            margin: 0 auto;
            background: white;
            border-radius: 20px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            overflow: hidden;
        }}
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px 20px;
            text-align: center;
        }}
        .header h1 {{
            margin: 0;
            font-size: 1.75rem;
            font-weight: bold;
            line-height: 1.3;
        }}
        .header .subtitle {{
            margin-top: 8px;
            font-size: 0.95rem;
            opacity: 0.9;
        }}
        .info-paciente {{
            background: #f8f9fa;
            padding: 20px 15px;
            border-left: 5px solid #667eea;
            margin: 20px 15px;
            border-radius: 10px;
        }}
        .info-paciente h3 {{
            color: #667eea;
            margin: 0 0 12px 0;
            font-size: 1.1rem;
        }}
        .info-item {{
            display: block;
            margin-bottom: 8px;
            font-size: 0.95rem;
            line-height: 1.5;
        }}
        .info-label {{
            font-weight: bold;
            color: #495057;
            display: block;
            margin-bottom: 3px;
        }}
        .info-value {{
            display: block;
            padding-left: 5px;
        }}
        .alert-summary {{
            text-align: center;
            padding: 20px 15px;
            background: #fff3cd;
            border-left: 5px solid #ffc107;
            margin: 20px 15px;
            border-radius: 10px;
        }}
        .alert-summary h2 {{
            color: #856404;
            margin: 0;
            font-size: 1.5rem;
            line-height: 1.3;
        }}
        .alert-summary p {{
            margin: 8px 0 0 0;
            color: #856404;
            font-size: 0.95rem;
        }}
        .medicamentos-container {{
            padding: 20px 15px;
        }}
        .section-title {{
            font-size: 1.2rem;
            font-weight: bold;
            margin-bottom: 15px;
            padding-bottom: 8px;
            border-bottom: 3px solid;
            line-height: 1.3;
        }}
        .section-hoy .section-title {{
            color: #dc3545;
            border-color: #dc3545;
        }}
        .section-manana .section-title {{
            color: #fd7e14;
            border-color: #fd7e14;
        }}
        .section-proximas .section-title {{
            color: #ffc107;
            border-color: #ffc107;
        }}
        .medicamento-card {{
            background: white;
            border-radius: 12px;
            padding: 15px;
            margin-bottom: 15px;
            box-shadow: 0 3px 10px rgba(0,0,0,0.1);
            border-left: 5px solid;
        }}
        .card-hoy {{
            border-color: #dc3545;
            background: linear-gradient(to right, #ffebee, white);
        }}
        .card-manana {{
            border-color: #fd7e14;
            background: linear-gradient(to right, #fff3e0, white);
        }}
        .card-proxima {{
            border-color: #ffc107;
            background: linear-gradient(to right, #fffde7, white);
        }}
        .medicamento-nombre {{
            font-size: 1.2rem;
            font-weight: bold;
            margin-bottom: 8px;
            color: #212529;
            line-height: 1.3;
            word-wrap: break-word;
        }}
        .medicamento-uso {{
            font-size: 0.9rem;
            color: #6c757d;
            margin-bottom: 12px;
            font-style: italic;
            line-height: 1.4;
        }}
        .medicamento-info {{
            display: block;
        }}
        .fecha-revision {{
            display: block;
            margin-bottom: 10px;
            font-size: 0.9rem;
            color: #495057;
        }}
        .fecha-revision strong {{
            display: inline-block;
            margin-right: 5px;
        }}
        .badge-dias {{
            display: inline-block;
            font-size: 0.85rem;
            padding: 6px 12px;
            border-radius: 50px;
            font-weight: bold;
            margin-top: 5px;
            text-align: center;
        }}
        .badge-hoy {{
            background: #dc3545;
            color: white;
        }}
        .badge-manana {{
            background: #fd7e14;
            color: white;
        }}
        .badge-proxima {{
            background: #ffc107;
            color: #000;
        }}
        .footer {{
            background: #f8f9fa;
            padding: 20px 15px;
            text-align: center;
            color: #6c757d;
            font-size: 0.85rem;
        }}
        .footer hr {{
            margin: 15px 0;
            border-color: #dee2e6;
        }}
        .footer p {{
            margin: 8px 0;
            line-height: 1.5;
        }}
        
        /* Media queries para pantallas muy pequeñas */
        @media (max-width: 480px) {{
            body {{
                padding: 5px;
            }}
            .header {{
                padding: 20px 15px;
            }}
            .header h1 {{
                font-size: 1.5rem;
            }}
            .header .subtitle {{
                font-size: 0.85rem;
            }}
            .info-paciente,
            .alert-summary,
            .medicamentos-container {{
                margin: 15px 10px;
                padding: 15px 12px;
            }}
            .medicamento-card {{
                padding: 12px;
            }}
            .medicamento-nombre {{
                font-size: 1.1rem;
            }}
            .section-title {{
                font-size: 1.1rem;
            }}
            .alert-summary h2 {{
                font-size: 1.3rem;
            }}
        }}
        
        /* Para clientes de email (Gmail, Outlook) */
        @media screen and (max-width: 600px) {{
            .container-email {{
                border-radius: 10px !important;
            }}
            table {{
                width: 100% !important;
            }}
        }}
    </style>
</head>
<body>
    <div class="container-email">
        <!-- Header -->
        <div class="header">
            <h1>🏥 Sistema de Alertas de Medicamentos</h1>
            <div class="subtitle">Control y seguimiento automatizado</div>
        </div>
        
        <!-- Información del Paciente -->
        <div class="info-paciente">
            <h3>📋 Información del Paciente</h3>
            <div class="info-item">
                <span class="info-label">👤 Paciente:</span>
                <span class="info-value">{info_paciente['paciente']}</span>
            </div>
            <div class="info-item">
                <span class="info-label">📍 Ubicación:</span>
                <span class="info-value">{info_paciente['ubicacion']}</span>
            </div>
        </div>
        
        <!-- Resumen de Alertas -->
        <div class="alert-summary">
            <h2>⚠️ {num_alertas} Medicamentos Requieren Atención</h2>
            <p>Fechas de revisión próximas en los siguientes {DIAS_ALERTA} días</p>
        </div>
        
        <!-- Medicamentos -->
        <div class="medicamentos-container">
    """
    
    # Alertas para HOY
    if alertas_hoy:
        html += """
            <div class="section-hoy">
                <div class="section-title">🔴 URGENTE - Revisión HOY</div>
        """
        for alerta in alertas_hoy:
            html += f"""
                <div class="medicamento-card card-hoy">
                    <div class="medicamento-nombre">{alerta['medicamento']}</div>
                    <div class="medicamento-uso">💊 {alerta['uso']}</div>
                    <div class="medicamento-info">
                        <div class="fecha-revision">
                            📅 <strong>Revisión:</strong> {alerta['fecha'].strftime('%d/%m/%Y')}
                        </div>
                        <span class="badge-dias badge-hoy">⏰ HOY - Acción Inmediata</span>
                    </div>
                </div>
            """
        html += "</div>"
    
    # Alertas para MAÑANA
    if alertas_manana:
        html += """
            <div class="section-manana">
                <div class="section-title">🟠 IMPORTANTE - Revisión MAÑANA</div>
        """
        for alerta in alertas_manana:
            html += f"""
                <div class="medicamento-card card-manana">
                    <div class="medicamento-nombre">{alerta['medicamento']}</div>
                    <div class="medicamento-uso">💊 {alerta['uso']}</div>
                    <div class="medicamento-info">
                        <div class="fecha-revision">
                            📅 <strong>Revisión:</strong> {alerta['fecha'].strftime('%d/%m/%Y')}
                        </div>
                        <span class="badge-dias badge-manana">⏰ 1 día restante</span>
                    </div>
                </div>
            """
        html += "</div>"
    
    # Alertas PRÓXIMAS
    if alertas_proximas:
        html += """
            <div class="section-proximas">
                <div class="section-title">🟡 PRÓXIMAMENTE - Planificar Revisión</div>
        """
        for alerta in alertas_proximas:
            html += f"""
                <div class="medicamento-card card-proxima">
                    <div class="medicamento-nombre">{alerta['medicamento']}</div>
                    <div class="medicamento-uso">💊 {alerta['uso']}</div>
                    <div class="medicamento-info">
                        <div class="fecha-revision">
                            📅 <strong>Revisión:</strong> {alerta['fecha'].strftime('%d/%m/%Y')}
                        </div>
                        <span class="badge-dias badge-proxima">⏰ {alerta['dias_restantes']} días restantes</span>
                    </div>
                </div>
            """
        html += "</div>"
    
    html += f"""
        </div>
        
        <!-- Footer -->
        <div class="footer">
            <hr>
            <p><strong>🕐 Revisión realizada:</strong> {fecha_revision}</p>
            <p><strong>📧 Sistema automatizado</strong> - GitHub Actions</p>
            <p><strong>☁️ Ejecutado en la nube</strong></p>
            <p style="margin-top: 20px; font-size: 0.85rem;">
                Este email fue generado automáticamente por el sistema de control de medicamentos.
            </p>
        </div>
    </div>
</body>
</html>
    """
    
    return html

def enviar_whatsapp(telefono, mensaje, info_paciente):
    """Envía mensaje por WhatsApp usando CallMeBot API (gratis)"""
    try:
        if not telefono:
            log("⚠️ No se configuró número de WhatsApp")
            return False
        
        log(f"Intentando enviar WhatsApp a: {telefono}")
        
        # CallMeBot API (gratis, sin registro previo)
        # NOTA: El número debe estar registrado en CallMeBot primero
        # Más info: https://www.callmebot.com/blog/free-api-whatsapp-messages/
        
        # Mensaje simplificado para WhatsApp
        texto = f"""
🏥 *ALERTA DE MEDICAMENTOS*

👤 Paciente: {info_paciente['paciente']}
📍 Ubicación: {info_paciente['ubicacion']}

{mensaje}

🤖 Sistema automatizado
        """.strip()
        
        # URL de la API de CallMeBot
        url = "https://api.callmebot.com/whatsapp.php"
        
        params = {
            'phone': telefono,
            'text': texto,
            'apikey': WHATSAPP_API_KEY
        }
        
        response = requests.get(url, params=params, timeout=10)
        
        if response.status_code == 200:
            log("✅ Mensaje de WhatsApp enviado correctamente")
            return True
        else:
            log(f"⚠️ Error al enviar WhatsApp: {response.status_code}")
            return False
            
    except Exception as e:
        log(f"⚠️ No se pudo enviar WhatsApp: {str(e)}")
        return False

def crear_mensaje_whatsapp(alertas):
    """Crea un mensaje resumido para WhatsApp"""
    mensaje = f"⚠️ *{len(alertas)} medicamentos* requieren revisión:\n\n"
    
    for i, alerta in enumerate(alertas[:5], 1):  # Máximo 5 para no saturar
        if alerta['dias_restantes'] == 0:
            urgencia = "🔴 HOY"
        elif alerta['dias_restantes'] == 1:
            urgencia = "🟠 MAÑANA"
        else:
            urgencia = f"🟡 {alerta['dias_restantes']} días"
        
        mensaje += f"{i}. *{alerta['medicamento']}*\n"
        mensaje += f"   {urgencia} - {alerta['fecha'].strftime('%d/%m/%Y')}\n\n"
    
    if len(alertas) > 5:
        mensaje += f"...y {len(alertas) - 5} más. Revisa tu email."
    
    return mensaje

def enviar_email(destinatario, asunto, cuerpo_html, archivo_adjunto=None):
    """Envía un email vía Gmail SMTP"""
    try:
        log("Preparando email...")
        
        mensaje = MIMEMultipart()
        mensaje['From'] = GMAIL_USUARIO
        mensaje['To'] = destinatario
        mensaje['Subject'] = asunto
        
        mensaje.attach(MIMEText(cuerpo_html, 'html', 'utf-8'))
        
        if archivo_adjunto and os.path.exists(archivo_adjunto):
            log(f"Adjuntando archivo: {archivo_adjunto}")
            with open(archivo_adjunto, 'rb') as archivo:
                parte = MIMEBase('application', 'octet-stream')
                parte.set_payload(archivo.read())
                encoders.encode_base64(parte)
                nombre_archivo = os.path.basename(archivo_adjunto)
                parte.add_header('Content-Disposition', f'attachment; filename= {nombre_archivo}')
                mensaje.attach(parte)
        
        log("Conectando con Gmail SMTP...")
        servidor = smtplib.SMTP('smtp.gmail.com', 587)
        servidor.starttls()
        
        log("Autenticando...")
        servidor.login(GMAIL_USUARIO, GMAIL_PASSWORD)
        
        log("Enviando email...")
        texto = mensaje.as_string()
        servidor.sendmail(GMAIL_USUARIO, destinatario, texto)
        servidor.quit()
        
        log("✅ Email enviado exitosamente!")
        return True
    
    except Exception as e:
        log(f"❌ ERROR al enviar email: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """Función principal"""
    log("="*70)
    log("SISTEMA DE REVISIÓN AUTOMÁTICA - VERSIÓN MEJORADA")
    log("="*70)
    
    if not all([GMAIL_USUARIO, GMAIL_PASSWORD, EMAIL_DESTINO]):
        log("❌ ERROR: Faltan variables de entorno")
        sys.exit(1)
    
    if not os.path.exists(RUTA_EXCEL):
        log(f"❌ ERROR: No se encontró el archivo Excel: {RUTA_EXCEL}")
        sys.exit(1)
    
    # Buscar alertas
    alertas, info_paciente = leer_excel_y_buscar_alertas(RUTA_EXCEL)
    
    if alertas is None:
        log("❌ No se pudo leer el archivo Excel")
        sys.exit(1)
    
    if len(alertas) > 0:
        log(f"\n🚨 Se encontraron {len(alertas)} alertas. Preparando notificaciones...")
        
        # Crear email HTML con Bootstrap
        cuerpo_html = crear_html_email_bootstrap(alertas, info_paciente)
        asunto = f"🏥 ALERTAS: {len(alertas)} Medicamentos - {info_paciente['paciente']}"
        
        # Enviar email
        resultado_email = enviar_email(EMAIL_DESTINO, asunto, cuerpo_html, RUTA_EXCEL)
        
        if resultado_email:
            log(f"✅ Email enviado a: {EMAIL_DESTINO}")
        else:
            log("❌ El email no pudo ser enviado")
        
        # Enviar WhatsApp si hay número configurado
        if info_paciente['telefono']:
            mensaje_wa = crear_mensaje_whatsapp(alertas)
            enviar_whatsapp(info_paciente['telefono'], mensaje_wa, info_paciente)
        else:
            log("ℹ️ No se envió WhatsApp (número no configurado en celda I4)")
        
        log(f"✅ Proceso completado")
    else:
        log("✅ No se encontraron alertas. No se envió ninguna notificación.")
    
    log("="*70)
    log("PROCESO FINALIZADO")
    log("="*70)

if __name__ == "__main__":
    main()
