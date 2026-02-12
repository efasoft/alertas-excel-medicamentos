"""
SISTEMA DE ALERTAS DE MEDICAMENTOS
Versión con diseño personalizado según PDF
Lee datos desde Google Drive
"""

import openpyxl
from datetime import datetime, date
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os
import sys
import requests

# Configuración desde variables de entorno
GMAIL_USUARIO = os.environ.get('GMAIL_USUARIO')
GMAIL_PASSWORD = os.environ.get('GMAIL_PASSWORD')
EMAIL_DESTINO = os.environ.get('EMAIL_DESTINO')
WHATSAPP_API_KEY = os.environ.get('WHATSAPP_API_KEY', '')

# Archivo Excel
RUTA_EXCEL = "medicamentos_alertas.xlsx"

# Configuración
DIAS_ALERTA = 3
FILA_INICIO = 18

def log(mensaje):
    """Registrar mensajes con timestamp"""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{timestamp}] {mensaje}")

def extraer_imagen_paciente(ruta_excel):
    """Extrae la imagen del paciente del Excel y la convierte a base64"""
    try:
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from PIL import Image
        import io
        import base64
        
        workbook = openpyxl.load_workbook(ruta_excel)
        sheet = workbook.active
        
        log("Buscando imagen del paciente en el Excel...")
        
        # Buscar imágenes en la hoja
        if not hasattr(sheet, '_images'):
            log("No se encontró el atributo _images en la hoja")
            workbook.close()
            return None
            
        imagenes = sheet._images
        log(f"Total de imágenes encontradas: {len(imagenes)}")
        
        for idx, image in enumerate(imagenes, 1):
            # La imagen del paciente debería estar en la zona L-M (columnas 11-12-13)
            if hasattr(image, 'anchor') and hasattr(image.anchor, '_from'):
                col = image.anchor._from.col
                row = image.anchor._from.row
                log(f"Imagen #{idx}: Columna {col} ({chr(65 + col) if col < 26 else 'Z+'}), Fila {row}")
                
                # Si está en las columnas L o M (11, 12, o 13) y filas 4-12
                if 11 <= col <= 13 and 4 <= row <= 12:
                    log(f"✓ Imagen encontrada en la zona esperada!")
                    # Convertir a base64
                    img_data = image._data()
                    img = Image.open(io.BytesIO(img_data))
                    
                    # Redimensionar si es muy grande
                    img.thumbnail((200, 200), Image.Resampling.LANCZOS)
                    
                    # Convertir a base64
                    buffered = io.BytesIO()
                    img.save(buffered, format="PNG")
                    img_base64 = base64.b64encode(buffered.getvalue()).decode()
                    
                    workbook.close()
                    log("✓ Imagen del paciente extraída correctamente")
                    return f"data:image/png;base64,{img_base64}"
        
        workbook.close()
        log("⚠ No se encontró imagen en la zona L-M, filas 5-12")
        return None
    except Exception as e:
        log(f"Error al extraer la imagen del paciente: {e}")
        import traceback
        traceback.print_exc()
        return None

def leer_info_paciente(sheet):
    """Lee la información del paciente desde las celdas específicas"""
    try:
        paciente = sheet['B5'].value or "No especificado"
        responsable = sheet['B9'].value or "No especificado"
        telefono_whatsapp = sheet['I9'].value or ""
        
        if telefono_whatsapp:
            telefono_whatsapp = str(telefono_whatsapp).replace(" ", "").replace("-", "").replace("+", "")
        
        return {
            'paciente': paciente,
            'responsable': responsable,
            'telefono': telefono_whatsapp
        }
    except Exception as e:
        log(f"Error al leer información del paciente: {e}")
        return {
            'paciente': "No especificado",
            'responsable': "No especificado",
            'telefono': ""
        }

def leer_excel_y_buscar_alertas(ruta_archivo):
    """Lee el archivo Excel y busca fechas próximas en columna J desde fila 18"""
    try:
        log(f"Abriendo archivo Excel: {ruta_archivo}")
        workbook = openpyxl.load_workbook(ruta_archivo, data_only=True)
        sheet = workbook.active
        
        info_paciente = leer_info_paciente(sheet)
        log(f"Paciente: {info_paciente['paciente']}")
        log(f"Responsable: {info_paciente['responsable']}")
        
        # Extraer imagen del paciente
        imagen_paciente = extraer_imagen_paciente(ruta_archivo)
        info_paciente['imagen'] = imagen_paciente
        
        alertas = []
        fecha_hoy = date.today()
        columna_fecha = 10
        
        log(f"Revisando columna J desde fila {FILA_INICIO}")
        log(f"Buscando fechas con menos de {DIAS_ALERTA} días...")
        
        for fila in range(FILA_INICIO, sheet.max_row + 1):
            celda = sheet.cell(row=fila, column=columna_fecha)
            valor = celda.value
            
            if isinstance(valor, datetime):
                fecha_celda = valor.date()
                dias_restantes = (fecha_celda - fecha_hoy).days
                
                if 0 <= dias_restantes < DIAS_ALERTA:
                    nombre_medicamento = sheet.cell(row=fila, column=1).value or "Medicamento sin nombre"
                    uso_medicamento = sheet.cell(row=fila, column=2).value or "Uso no especificado"
                    
                    alerta = {
                        'fila': fila,
                        'fecha': fecha_celda,
                        'dias_restantes': dias_restantes,
                        'medicamento': str(nombre_medicamento),
                        'uso': str(uso_medicamento)
                    }
                    alertas.append(alerta)
                    log(f"  ⚠️ Alerta: {nombre_medicamento} - Fila {fila}, Fecha: {fecha_celda}, Días: {dias_restantes}")
        
        workbook.close()
        log(f"Total de alertas encontradas: {len(alertas)}")
        return alertas, info_paciente
    
    except FileNotFoundError:
        log(f"❌ ERROR: No se encontró el archivo: {ruta_archivo}")
        return None, None
    except Exception as e:
        log(f"❌ ERROR al leer Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return None, None

def crear_html_email_personalizado(alertas, info_paciente):
    """Crea email HTML con diseño moderno glassmorphism"""
    num_alertas = len(alertas)
    fecha_revision = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    
    # Diccionarios para traducir meses y días
    meses_es = {
        1: 'ENE', 2: 'FEB', 3: 'MAR', 4: 'ABR', 5: 'MAY', 6: 'JUN',
        7: 'JUL', 8: 'AGO', 9: 'SEP', 10: 'OCT', 11: 'NOV', 12: 'DIC'
    }
    
    dias_es = {
        0: 'LUN', 1: 'MAR', 2: 'MIÉ', 3: 'JUE', 4: 'VIE', 5: 'SÁB', 6: 'DOM'
    }
    
    # Foto del paciente (base64 o placeholder)
    foto_paciente = info_paciente.get('imagen') or 'data:image/svg+xml,<svg xmlns="http://www.w3.org/2000/svg" width="200" height="200"><rect fill="%23e0e0e0" width="200" height="200"/><text x="50%" y="50%" font-size="80" text-anchor="middle" dy=".3em">👤</text></svg>'
    
    html = f"""
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Control de Medicamentos</title>
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link href="https://fonts.googleapis.com/css2?family=Montserrat:wght@300;400;600;700;800&family=Raleway:wght@300;400;500;600;700&display=swap" rel="stylesheet">
    <style>
        * {{ 
            box-sizing: border-box; 
            margin: 0; 
            padding: 0; 
        }}
        body {{ 
            font-family: 'Raleway', sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 50%, #f093fb 100%); 
            padding: 30px 20px;
            min-height: 100vh;
        }}
        .container {{ 
            max-width: 1000px; 
            margin: 0 auto; 
            background: rgba(255, 255, 255, 0.98);
            border-radius: 30px; 
            overflow: hidden; 
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
        }}
        
        /* Header con imagen de píldoras */
        .header {{ 
            background: #667eea;
            color: white; 
            padding: 50px 40px; 
            text-align: center; 
            position: relative;
        }}
        .header h1 {{ 
            font-family: 'Montserrat', sans-serif;
            font-size: 2.5rem; 
            font-weight: 800; 
            color: #ffffff;
            text-shadow: 0 4px 12px rgba(0,0,0,0.3);
            letter-spacing: 3px;
            text-transform: uppercase;
        }}
        
        /* Contenedor de tarjetas apiladas */
        .info-cards {{ 
            padding: 50px 40px;
            background: #f8f9fa;
        }}
        
        /* Tarjeta del paciente (verde esmeralda sólido) */
        .card-paciente {{ 
            background: #059669;
            color: white;
            border-radius: 25px;
            padding: 40px;
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 35px;
            box-shadow: 0 15px 35px rgba(5, 150, 105, 0.3);
            margin-bottom: 60px;
        }}
        .card-paciente .foto {{ 
            width: 140px; 
            height: 140px; 
            border-radius: 50%; 
            background: white;
            overflow: hidden;
            border: 5px solid rgba(255, 255, 255, 0.5);
            flex-shrink: 0;
            box-shadow: 0 8px 20px rgba(0,0,0,0.2);
            display: flex;
            align-items: center;
            justify-content: center;
        }}
        .card-paciente .foto img {{ 
            width: 100%; 
            height: 100%; 
            object-fit: cover; 
            display: block;
        }}
        .card-paciente .info {{ 
            flex: 1;
            text-align: center;
        }}
        .card-paciente .label {{ 
            font-family: 'Montserrat', sans-serif;
            font-size: 0.95rem; 
            font-weight: 700; 
            color: #fde047;
            margin-bottom: 12px;
            text-transform: uppercase;
            letter-spacing: 3px;
        }}
        .card-paciente .valor {{ 
            font-family: 'Montserrat', sans-serif;
            font-size: 2.2rem; 
            font-weight: 700; 
            color: #ffffff;
            line-height: 1.2;
        }}
        
        /* Tarjeta del responsable (azul índigo sólido) */
        .card-responsable {{ 
            background: #4f46e5;
            color: white;
            border-radius: 25px;
            padding: 40px;
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            text-align: center;
            gap: 20px;
            box-shadow: 0 15px 35px rgba(79, 70, 229, 0.3);
        }}
        .card-responsable .seccion {{
            display: flex;
            flex-direction: column;
            gap: 8px;
        }}
        .card-responsable .label {{ 
            font-family: 'Montserrat', sans-serif;
            font-size: 0.95rem; 
            font-weight: 700; 
            color: #ffffff;
            text-transform: uppercase;
            letter-spacing: 3px;
        }}
        .card-responsable .valor {{ 
            font-family: 'Montserrat', sans-serif;
            font-size: 2rem; 
            font-weight: 700; 
            color: #ffffff;
            line-height: 1.2;
        }}
        .card-responsable .telefono {{ 
            font-family: 'Raleway', sans-serif;
            font-size: 2rem; 
            font-weight: 700;
            color: #ffffff;
            text-align: center;
        }}
        
        /* Banner amarillo de advertencia */
        .alert-banner {{ 
            background: #fbbf24;
            color: #1f2937;
            padding: 35px 45px;
            margin: 0 40px 40px 40px;
            border-radius: 25px;
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            text-align: center;
            gap: 15px;
            box-shadow: 0 12px 30px rgba(251, 191, 36, 0.3);
        }}
        .alert-banner .icon {{ 
            font-size: 4rem;
            filter: drop-shadow(2px 2px 4px rgba(0,0,0,0.1));
        }}
        .alert-banner .texto {{ 
            font-family: 'Montserrat', sans-serif;
            font-size: 1.5rem;
            font-weight: 700;
            line-height: 1.5;
            text-transform: uppercase;
            letter-spacing: 1px;
            color: #1f2937;
        }}
        
        /* Container de medicamentos */
        .medicamentos-container {{ 
            padding: 0 40px 50px 40px; 
        }}
        
        /* Tarjeta de medicamento */
        .medicamento-card {{ 
            background: white;
            border-radius: 25px;
            margin-bottom: 30px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.08);
            overflow: hidden;
            display: flex;
            border: 1px solid rgba(0,0,0,0.05);
        }}
        
        /* Calendario lateral (rojo-carmesí) */
        .calendario {{ 
            background: #dc2626;
            color: white;
            width: 160px;
            padding: 25px;
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            text-align: center;
            flex-shrink: 0;
        }}
        .calendario .dia-semana {{ 
            font-family: 'Montserrat', sans-serif;
            font-size: 1.1rem; 
            font-weight: 700; 
            margin-bottom: 8px;
            letter-spacing: 2px;
            color: #ffffff;
        }}
        .calendario .dia {{ 
            font-family: 'Montserrat', sans-serif;
            font-size: 4.5rem; 
            font-weight: 800; 
            line-height: 1;
            margin-bottom: 8px;
            color: #ffffff;
        }}
        .calendario .mes {{ 
            font-family: 'Montserrat', sans-serif;
            font-size: 1.3rem; 
            font-weight: 700; 
            letter-spacing: 2px;
            color: #ffffff;
        }}
        
        /* Contenido del medicamento */
        .medicamento-contenido {{ 
            flex: 1;
            padding: 35px 40px;
            display: flex;
            flex-direction: column;
            gap: 18px;
        }}
        .medicamento-nombre {{ 
            font-family: 'Montserrat', sans-serif;
            font-size: 2rem; 
            font-weight: 700; 
            color: #1f2937;
            line-height: 1.2;
        }}
        .medicamento-uso {{ 
            font-family: 'Raleway', sans-serif;
            font-size: 1.1rem; 
            color: #6b7280;
            font-weight: 400;
        }}
        
        /* Badge de días restantes */
        .badge-dias {{ 
            display: inline-block;
            background: #f97316;
            color: white;
            padding: 10px 20px;
            border-radius: 50px;
            font-family: 'Montserrat', sans-serif;
            font-size: 0.95rem;
            font-weight: 700;
            margin-top: 12px;
            box-shadow: 0 4px 12px rgba(249, 115, 22, 0.3);
            text-align: center;
            text-transform: uppercase;
        }}
        
        /* Footer */
        .footer {{ 
            background: #1e293b;
            color: white;
            padding: 40px;
            text-align: center;
        }}
        .footer-info {{ 
            display: block;
            margin-bottom: 20px;
            font-family: 'Raleway', sans-serif;
            font-size: 0.95rem;
            font-weight: 400;
            color: #ffffff;
            line-height: 2;
        }}
        .footer-info div {{
            color: #ffffff;
            display: inline;
            margin: 0 15px;
        }}
        .footer-info div:after {{
            content: " | ";
            margin-left: 15px;
        }}
        .footer-info div:last-child:after {{
            content: "";
        }}
        .footer p {{
            font-family: 'Raleway', sans-serif;
            font-size: 0.9rem;
            opacity: 0.7;
            font-weight: 300;
            color: #ffffff;
        }}
        
        /* Responsive */
        @media (max-width: 768px) {{
            body {{ padding: 15px; }}
            .header {{ padding: 35px 20px; }}
            .header h1 {{ font-size: 1.6rem; letter-spacing: 1px; }}
            .info-cards {{ padding: 30px 20px; gap: 25px; }}
            .card-paciente, .card-responsable {{ 
                padding: 25px; 
                flex-direction: column;
            }}
            .card-paciente {{ margin-bottom: 30px; }}
            .card-paciente .foto {{ width: 100px; height: 100px; }}
            .card-paciente .label {{ font-size: 0.75rem; }}
            .card-paciente .valor {{ font-size: 1.5rem; }}
            .card-responsable .label {{ font-size: 0.75rem; }}
            .card-responsable .valor {{ font-size: 1.4rem; color: #ffffff !important; }}
            .card-responsable .telefono {{ font-size: 1.4rem; color: #ffffff !important; }}
            .alert-banner {{ 
                margin: 0 20px 30px 20px;
                padding: 25px;
                text-align: center;
                gap: 10px;
            }}
            .alert-banner .icon {{ font-size: 2.5rem; }}
            .alert-banner .texto {{ font-size: 1rem; letter-spacing: 0.5px; }}
            .medicamentos-container {{ padding: 0 20px 30px 20px; }}
            .medicamento-card {{ flex-direction: column; margin-bottom: 20px; }}
            .calendario {{ width: 100%; padding: 15px; }}
            .calendario .dia-semana {{ font-size: 0.9rem; }}
            .calendario .dia {{ font-size: 3rem; }}
            .calendario .mes {{ font-size: 1rem; }}
            .medicamento-contenido {{ padding: 20px; }}
            .medicamento-nombre {{ font-size: 1.4rem; }}
            .medicamento-uso {{ font-size: 0.9rem; }}
            .badge-dias {{ font-size: 0.85rem; padding: 8px 16px; }}
            .footer {{ padding: 25px 20px; }}
            .footer-info {{ font-size: 0.85rem; line-height: 2.5; }}
            .footer-info div {{ display: block; margin: 5px 0; }}
            .footer-info div:after {{ content: ""; }}
            .footer p {{ font-size: 0.8rem; }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <!-- Header con imagen de píldoras -->
        <div class="header">
            <h1>CONTROL DE MEDICAMENTOS</h1>
        </div>
        
        <!-- Tarjetas apiladas de información -->
        <div class="info-cards">
            <!-- Tarjeta verde del paciente -->
            <div class="card-paciente">
                <div class="foto">
                    <img src="{foto_paciente}" alt="Foto paciente">
                </div>
                <div class="info">
                    <div class="label">PACIENTE</div>
                    <div class="valor">{info_paciente['paciente']}</div>
                </div>
            </div>
            
            <!-- Tarjeta azul del responsable -->
            <div class="card-responsable">
                <div class="seccion">
                    <div class="label">RESPONSABLE</div>
                    <div class="valor">{info_paciente['responsable']}</div>
                </div>
                <div class="telefono">
                    {info_paciente['telefono'] or 'Sin teléfono'}
                </div>
            </div>
        </div>
        
        <!-- Banner amarillo de advertencia -->
        <div class="alert-banner">
            <div class="icon">✋</div>
            <div class="texto">
                Medicamentos que están próximos a agotarse y requieren atención
            </div>
        </div>
        
        <!-- Lista de medicamentos -->
        <div class="medicamentos-container">
"""
    
    # Generar tarjetas de medicamentos
    for alerta in alertas:
        fecha = alerta['fecha']
        dia_semana = dias_es[fecha.weekday()]
        dia = fecha.day
        mes = meses_es[fecha.month]
        dias_texto = f"Quedan {alerta['dias_restantes']:02d} días" if alerta['dias_restantes'] > 0 else "VENCE HOY"
        
        html += f"""
            <div class="medicamento-card">
                <!-- Calendario lateral -->
                <div class="calendario">
                    <div class="dia-semana">{dia_semana}</div>
                    <div class="dia">{dia}</div>
                    <div class="mes">{mes}</div>
                </div>
                
                <!-- Contenido del medicamento -->
                <div class="medicamento-contenido">
                    <div class="medicamento-nombre">{alerta['medicamento']}</div>
                    <div class="medicamento-uso">{alerta['uso']}</div>
                    <div class="badge-dias">{dias_texto}</div>
                </div>
            </div>
        """
    
    html += f"""
        </div>
        
        <!-- Footer -->
        <div class="footer">
            <div class="footer-info">
                <div>Revisión: {fecha_revision}</div>
                <div>Sistema Automatizado</div>
                <div>Desarrollado por: Ernesto Fernandez +34 611131467</div>
            </div>
            <p style="margin-top: 20px;">
                Este correo fue generado automáticamente por el sistema de alertas de medicamentos
            </p>
        </div>
    </div>
</body>
</html>
    """
    
    return html

def enviar_whatsapp(telefono, mensaje, info_paciente):
    """Envía mensaje por WhatsApp"""
    try:
        if not telefono:
            return False
        
        texto = f"🏥 ALERTA MEDICAMENTOS\n👤 {info_paciente['paciente']}\n👨‍⚕️ {info_paciente['responsable']}\n\n{mensaje}"
        url = "https://api.callmebot.com/whatsapp.php"
        params = {'phone': telefono, 'text': texto, 'apikey': WHATSAPP_API_KEY}
        response = requests.get(url, params=params, timeout=10)
        return response.status_code == 200
    except:
        return False

def crear_mensaje_whatsapp(alertas):
    """Crea mensaje resumido para WhatsApp"""
    mensaje = f"⚠️ {len(alertas)} medicamentos requieren revisión:\n\n"
    for i, alerta in enumerate(alertas[:5], 1):
        urgencia = "🔴 HOY" if alerta['dias_restantes'] == 0 else f"🟡 {alerta['dias_restantes']} días"
        mensaje += f"{i}. {alerta['medicamento']}\n   {urgencia} - {alerta['fecha'].strftime('%d/%m/%Y')}\n\n"
    if len(alertas) > 5:
        mensaje += f"...y {len(alertas) - 5} más."
    return mensaje

def enviar_email(destinatario, asunto, cuerpo_html, archivo_adjunto=None):
    """Envía email vía Gmail SMTP"""
    try:
        log("Preparando email...")
        mensaje = MIMEMultipart()
        mensaje['From'] = GMAIL_USUARIO
        mensaje['To'] = destinatario
        mensaje['Subject'] = asunto
        mensaje.attach(MIMEText(cuerpo_html, 'html', 'utf-8'))
        
        if archivo_adjunto and os.path.exists(archivo_adjunto):
            with open(archivo_adjunto, 'rb') as archivo:
                parte = MIMEBase('application', 'octet-stream')
                parte.set_payload(archivo.read())
                encoders.encode_base64(parte)
                parte.add_header('Content-Disposition', f'attachment; filename= {os.path.basename(archivo_adjunto)}')
                mensaje.attach(parte)
        
        log("Conectando con Gmail...")
        servidor = smtplib.SMTP('smtp.gmail.com', 587)
        servidor.starttls()
        servidor.login(GMAIL_USUARIO, GMAIL_PASSWORD)
        servidor.sendmail(GMAIL_USUARIO, destinatario, mensaje.as_string())
        servidor.quit()
        
        log("✅ Email enviado exitosamente!")
        return True
    except Exception as e:
        log(f"❌ ERROR al enviar email: {str(e)}")
        return False

def main():
    """Función principal"""
    log("="*70)
    log("SISTEMA DE ALERTAS DE MEDICAMENTOS - VERSIÓN PERSONALIZADA")
    log("="*70)
    
    if not all([GMAIL_USUARIO, GMAIL_PASSWORD, EMAIL_DESTINO]):
        log("❌ ERROR: Faltan variables de entorno")
        sys.exit(1)
    
    if not os.path.exists(RUTA_EXCEL):
        log(f"❌ ERROR: No se encontró el archivo Excel: {RUTA_EXCEL}")
        sys.exit(1)
    
    alertas, info_paciente = leer_excel_y_buscar_alertas(RUTA_EXCEL)
    
    if alertas is None:
        log("❌ No se pudo leer el archivo Excel")
        sys.exit(1)
    
    if len(alertas) > 0:
        log(f"\n🚨 Se encontraron {len(alertas)} alertas")
        
        cuerpo_html = crear_html_email_personalizado(alertas, info_paciente)
        asunto = f"🏥 ALERTAS: {len(alertas)} Medicamentos - {info_paciente['paciente']}"
        
        if enviar_email(EMAIL_DESTINO, asunto, cuerpo_html, RUTA_EXCEL):
            log(f"✅ Email enviado a: {EMAIL_DESTINO}")
        
        if info_paciente['telefono']:
            mensaje_wa = crear_mensaje_whatsapp(alertas)
            enviar_whatsapp(info_paciente['telefono'], mensaje_wa, info_paciente)
    else:
        log("✅ No se encontraron alertas")
    
    log("="*70)
    log("PROCESO FINALIZADO")
    log("="*70)

if __name__ == "__main__":
    main()
